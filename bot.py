# -*- coding: utf-8 -*-

import logging
import asyncio
import aiohttp
import requests
import schedule
import time
import os
import openpyxl
import re
import sys
import pytz
import pandas as pd
from aiogram import F
from aiogram.filters import StateFilter
from aiogram import Bot, Dispatcher, Router, types
from aiogram.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# Настройка логирования и бота
logging.basicConfig(level=logging.INFO)
TOKEN = "7766027837:AAFFORwPFg_CCZ5iEx0saTzCQL-ihXoHvNA"
bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())
router = Router()

SCHEDULE_URL = "https://disk.yandex.ru/i/zBdSFy9HRBb9Pw"
FILE_PATH = "44.03.01 Информатика.xlsx"
TEMP_FILE_PATH = "temp.xlsx"
last_update_time = None
UPDATE_INFO_FILE = "last_update.txt"
UPDATE_TIME = "22:00"
ADMIN_ID = 916756380
UPDATE_STATUS_FILE = "update_status.txt"

#Создает файл-флаг перед обновлением
def set_update_flag():
    with open(UPDATE_STATUS_FILE, "w") as file:
        file.write("updated")

#Проверяет, было ли обновление перед перезапуском
def check_update_flag():
    return os.path.exists(UPDATE_STATUS_FILE)

#Удаляет флаг обновления после перезапуска
def clear_update_flag():
    if os.path.exists(UPDATE_STATUS_FILE):
        os.remove(UPDATE_STATUS_FILE)

# Определение состояния для FSM
class FeedbackState(StatesGroup):
    waiting_for_feedback = State()

# Обработчик команды /feedback
@dp.message(Command("feedback"))
async def feedback_command(message: types.Message, state: FSMContext):
    markup = InlineKeyboardMarkup(inline_keyboard=[  # Исправленная разметка
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_feedback")]
    ])
    await message.answer(f"📝 Оставьте ваш отзыв, и мы обязательно его передадим администратору. Спасибо за обратную связь!\nЕсли передумали, просто нажмите кнопку ниже.", reply_markup=markup)
    await state.set_state(FeedbackState.waiting_for_feedback)

# Обработчик нажатия на кнопку отмены
@dp.callback_query(F.data == "cancel_feedback", StateFilter(FeedbackState.waiting_for_feedback))
async def cancel_feedback(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Ввод отзыва отменен. Если захотите оставить мнение позже, мы всегда рады его услышать!")
    user_data = await state.get_data()
    group = user_data.get("group")
    await state.clear()
    
    # Если группа сохранена, показываем пользователю соответствующие кнопки
    if group:
        await state.update_data(group=group)
        markup = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Сегодня"), KeyboardButton(text="Завтра")],
                [KeyboardButton(text="Неделя"), KeyboardButton(text="Выбрать дату")],
                [KeyboardButton(text="Следующая пара")]
            ],
            resize_keyboard=True
        )

# Обработчик текстового сообщения с отзывом
@dp.message(FeedbackState.waiting_for_feedback)
async def receive_feedback(message: types.Message, state: FSMContext):
    user = message.from_user
    feedback_text = message.text
    
    # Отправляем отзыв администратору
    admin_message = f"📩 Новый отзыв от @{user.username if user.username else user.first_name}:\n\n{feedback_text}"
    await bot.send_message(ADMIN_ID, admin_message)
    
    # Подтверждаем пользователю отправку
    await message.answer("✅ Ваш отзыв успешно отправлен! Спасибо за ваше мнение.")
    
     # Получаем сохраненную группу из контекста
    user_data = await state.get_data()
    group = user_data.get("group")  # Получаем сохраненную группу
    await state.clear()
    # Если группа сохранена, показываем пользователю соответствующие кнопки
    if group:
        await state.update_data(group=group)
        markup = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Сегодня"), KeyboardButton(text="Завтра")],
                [KeyboardButton(text="Неделя"), KeyboardButton(text="Выбрать дату")],
                [KeyboardButton(text="Следующая пара")]
            ],
            resize_keyboard=True
        )

# Хранение списка пользователей
def load_users():
    if os.path.exists("users.txt"):
        with open("users.txt", "r") as file:
            return set(int(line.strip()) for line in file)
    return set()

def save_user(user_id):
    if user_id not in subscribed_users:
        with open("users.txt", "a") as file:
            file.write(f"{user_id}\n")
        subscribed_users.add(user_id)

subscribed_users = load_users()

@dp.message(Command(commands=['subscribe']))
async def subscribe(message: types.Message):
    user_id = message.from_user.id
    if user_id not in subscribed_users:
        save_user(user_id)
        await message.answer("✅ Вы подписаны на уведомления об обновлении расписания!")
    else:
        await message.answer("📢 Вы уже подписаны на уведомления!")

# Функция загрузки даты последнего обновления
def load_last_update():
    if os.path.exists(UPDATE_INFO_FILE):
        with open(UPDATE_INFO_FILE, "r") as file:
            return file.read().strip()
    return "Неизвестно"

def save_last_update(timestamp):
    with open(UPDATE_INFO_FILE, "w") as file:
        file.write(timestamp)

last_update_time = load_last_update()

#Чтение даты последнего обновления из файла
def get_last_update_time():
    if os.path.exists("last_update.txt"):
        with open("last_update.txt", "r") as file:
            return file.read().strip()
    return "Неизвестно"

update_time = get_last_update_time()

#Получение прямую ссылку на файл с Яндекс.Диска
def get_direct_link(public_url):
    api_url = "https://cloud-api.yandex.net/v1/disk/public/resources/download"
    params = {"public_key": public_url}
    response = requests.get(api_url, params=params)
    if response.status_code == 200:
        return response.json().get("href")
    else:
        logging.error(f"Ошибка получения ссылки: {response.status_code}")
        return None

#Скачивание расписания с Яндекс.Диска с предварительным сравнением
async def download_schedule():
    global last_update_time
    direct_link = get_direct_link(SCHEDULE_URL)
    if not direct_link:
        return False

    # Переименовываем текущий файл в temp
    if os.path.exists(FILE_PATH):
        os.rename(FILE_PATH, TEMP_FILE_PATH)

    response = requests.get(direct_link)
    if response.status_code == 200:
        with open(FILE_PATH, "wb") as file:
            file.write(response.content)

        # Проверяем изменения в файлах
        if os.path.exists(TEMP_FILE_PATH):
            if compare_excel_files(TEMP_FILE_PATH, FILE_PATH):
                os.remove(FILE_PATH)  # Удаляем новый файл
                os.rename(TEMP_FILE_PATH, FILE_PATH)  # Возвращаем старый
                logging.info("Файл не изменился, обновление отменено.")
                return False
            else:
                os.remove(TEMP_FILE_PATH)  # Удаляем старый файл

        last_update_time = datetime.now().strftime('%d.%m.%Y %H:%M')
        save_last_update(last_update_time)
        logging.info(f"Файл {FILE_PATH} успешно обновлен.")

        await notify_users_after_update()  # Уведомляем пользователей
        return True
    else:
        logging.error(f"Ошибка загрузки файла: {response.status_code}")
        return False

#Скачивание расписания с Яндекс.Диска  
async def manual_download():
    global last_update_time
    direct_link = get_direct_link(SCHEDULE_URL)
    if not direct_link:
        return False

    response = requests.get(direct_link)
    if response.status_code == 200:
        with open(TEMP_FILE_PATH, "wb") as file:
            file.write(response.content)

        if os.path.exists(FILE_PATH):
            os.remove(FILE_PATH)
        os.rename(TEMP_FILE_PATH, FILE_PATH)

        last_update_time = datetime.now().strftime('%d.%m.%Y %H:%M')
        save_last_update(last_update_time)
        logging.info(f"Файл {FILE_PATH} успешно обновлен.")

        await notify_users_after_update() # Вызов уведомления после обновления
        return True
    else:
        logging.error(f"Ошибка загрузки файла: {response.status_code}")
        return False

#Сравнение столбцов E и F двух файлов
def compare_excel_files(file1, file2):
    try:
        df1 = pd.read_excel(file1, usecols=[4, 5])
        df2 = pd.read_excel(file2, usecols=[4, 5])
        return df1.equals(df2)
    except Exception as e:
        logging.error(f"Ошибка сравнения файлов: {e}")
        return False

#Оповещение пользователей о начале и окончании обновления
async def notify_users():
    for user_id in subscribed_users:
        try:
            await bot.send_message(user_id, "♻️ Пожалуйста, подождите, обновление расписания...")
        except Exception as e:
            logging.error(f"Ошибка при отправке сообщения пользователю {user_id}: {e}")

async def notify_users_after_update():  # Функция для уведомления после обновления
    await asyncio.sleep(5)  # Даем время на отправку
    for user_id in subscribed_users:
        try:
            await bot.send_message(user_id, "📅 Расписание обновлено! Нажмите /start для обновления данных.")
        except Exception as e:
            logging.error(f"Ошибка при отправке сообщения пользователю {user_id}: {e}")
            
#Запуск обновление и перезапуск бота автоматически
async def update_and_restart():
    success = await download_schedule()
    if success:
        set_update_flag()  # Устанавливаем флаг перед рестартом
        await notify_users()
        os.execv(sys.executable, [sys.executable] + sys.argv)  # Перезапуск кода
    else:
        logging.error("Обновление не удалось.")
        
#Запуск обновление и перезапуск бота вручную
async def manual_update_and_restart():
    await notify_users()
    success = await manual_download()
    if success:
        set_update_flag()  # Устанавливаем флаг перед рестартом
        os.execv(sys.executable, [sys.executable] + sys.argv)  # Перезапуск кода
    else:
        logging.error("Обновление не удалось.")

@dp.message(Command(commands=['update_schedule']))
async def manual_update(message: types.Message):
    await message.answer("⚙️ Обновляю расписание вручную...")
    last_update_time = datetime.now().strftime('%d.%m.%Y %H:%M')
    save_last_update(last_update_time)
    await manual_update_and_restart()
    
#Фоновая задача по проверке обновлений расписания
async def auto_update():
    while True:
        now = datetime.now().strftime("%H:%M")
        await asyncio.sleep(60)  # Проверка каждую минуту
        if now == UPDATE_TIME:
            logging.info("Начинаю автоматическое обновление...")
            await update_and_restart()
    
# Функция для подготовки данных из Excel
def unmerge_and_fill_cells(sheet):
    for merged_cell in list(sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
        top_left_value = sheet.cell(row=min_row, column=min_col).value  # Берём значение верхней левой ячейки
        sheet.unmerge_cells(str(merged_cell))  # Разъединяем ячейки

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_value  # Заполняем разъединённые ячейки

#Удаляет лишние пробелы и переносы строк в столбце B
def clean_column_b(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_col=2, max_col=2, min_row=2, values_only=False):
        cell = row[0]
        if cell.value:
            cell.value = str(cell.value).strip().replace("\n", "  ")

    wb.save("44.03.01 Информатика_unmerged.xlsx")

# Основной код
wb = load_workbook("44.03.01 Информатика.xlsx")
sheet = wb.active
unmerge_and_fill_cells(sheet)
wb.save("44.03.01 Информатика_unmerged.xlsx")

clean_column_b("44.03.01 Информатика_unmerged.xlsx")  # Очистка столбца B

# Открываем новый файл после обработки объединенных ячеек
def load_transformed_schedule(file_path="44.03.01 Информатика_unmerged.xlsx"):
    if 'openpyxl' in sys.modules:
        del sys.modules['openpyxl']  # Удаляем кеш openpyxl
    if 'pandas' in sys.modules:
        del sys.modules['pandas']  # Удаляем кеш pandas

    wb = load_workbook(file_path, data_only=True)  # Читаем заново
    ws = wb.active
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)
    return df

# Используем новый файл для работы в get_schedule
df = load_transformed_schedule()

# Функция получения расписания
def get_schedule(group, date):

    df = load_transformed_schedule()  # Загружаем актуальные данные
    
    schedule = []
    date_str = date.strftime("%Y-%m-%d")
    today = datetime.now().date()
    found_date = False
    practice_counter = 0  # Счётчик строк с практикой
    current_time = datetime.now().strftime("%H.%M")  # Получаем текущее время в формате ЧЧ.ММ

    for i in range(len(df)):
        cell_date = pd.to_datetime(df.iloc[i, 0], errors='coerce')
        if pd.notna(cell_date) and cell_date.strftime("%Y-%m-%d") == date_str:
            found_date = True
            for j in range(7):  # Проверяем 7 строк под текущей датой
                time_info = df.iloc[i + j, 1]  # Например: "1 пара 9.00-10.30"
                group_1_schedule = df.iloc[i + j, 4]
                group_2_schedule = df.iloc[i + j, 5]

                # Проверяем строки на наличие практики
                if group_1_schedule and "практика" in group_1_schedule.lower():
                    practice_counter += 1
                    continue

                if pd.notna(time_info):
                    # Используем регулярное выражение для выделения номера пары и времени
                    match = re.match(r"(\d+ пара)\s+(\d{1,2}\.\d{2}-\d{1,2}\.\d{2})", time_info)
                    if match:
                        pair_number = match.group(1)
                        pair_time = match.group(2)
                    else:
                        pair_number, pair_time = time_info, ""  # Если не удалось разделить корректно

                    pair_text = ""  # Инициализация переменной для текста пары

                    # Определяем текущую пару
                    is_current = date == today and is_current_pair(pair_time, current_time)
                    
                    # Формируем строку расписания
                    if pd.notna(group_1_schedule) and pd.notna(group_2_schedule) and group_1_schedule == group_2_schedule:
                        pair_text = f"📚{pair_number}📚\n{pair_time}\n🫂{group_1_schedule}\n"
                    elif group == 1 and pd.notna(group_1_schedule):
                        pair_text = f"📚{pair_number}📚\n{pair_time}\n{group_1_schedule}\n"
                    elif group == 2 and pd.notna(group_2_schedule):
                        pair_text = f"📚{pair_number}📚\n{pair_time}\n{group_2_schedule}\n"

                    # Добавляем выделение текущей пары
                    if is_current and pair_text:
                        pair_text = f"<b>{pair_text}</b>"

                    # Добавляем в расписание
                    if pair_text.strip(): 
                        schedule.append(pair_text)

            # Если нашлись строки с практикой — добавляем строку "Практика в школе"
            if practice_counter > 0:
                schedule.append("Практика в школе.\n")
        
            break

    return "".join(schedule) if found_date and schedule else "Нет занятий.\n"

# Функция для проверки, является ли пара текущей
def is_current_pair(pair_time, current_time):
    try:
        start_time_str, end_time_str = pair_time.split('-')
        start_time = datetime.strptime(start_time_str, "%H.%M")
        end_time = datetime.strptime(end_time_str, "%H.%M")

        # Сравниваем текущее время с временем начала и окончания пары
        current_time_dt = datetime.strptime(current_time, "%H.%M")
        if start_time <= current_time_dt <= end_time:
            return True
    except ValueError:
        return False
    return False

# Функция для поиска следующей пары
def get_next_class(group, date, current_time):
    schedule = ""
    date_str = date.strftime("%Y-%m-%d")
    next_class = "Сегодня больше нет занятий.\n"
    practice_counter = 0  # Счётчик строк с практикой

    for i in range(len(df)):
        cell_date = pd.to_datetime(df.iloc[i, 0], errors='coerce')
        if pd.notna(cell_date) and cell_date.strftime("%Y-%m-%d") == date_str:
            for j in range(7):
                time_info = df.iloc[i + j, 1]
                group_1_schedule = df.iloc[i + j, 4]
                group_2_schedule = df.iloc[i + j, 5]

                # Проверяем строки на наличие практики
                if group_1_schedule and "практика" in group_1_schedule.lower():
                    practice_counter += 1
                    continue
                try:
                    pair_start_time = re.search(r"(\d{1,2}\.\d{2})", time_info).group(1)
                    pair_start_time_dt = datetime.strptime(pair_start_time, "%H.%M")

                    if pd.notna(time_info):
                        match = re.match(r"(\d+ пара)\s+(\d{1,2}\.\d{2}-\d{1,2}\.\d{2})", time_info)
                        if match:
                            pair_number, pair_time = match.groups()
                        else:
                            pair_number, pair_time = time_info, ""
                    
                    if datetime.strptime(current_time, "%H.%M") < pair_start_time_dt:
                        if pd.notna(group_1_schedule) and group_1_schedule == group_2_schedule:
                            return f"📚{pair_number}📚\n{pair_time}\n🫂{group_1_schedule}\n"
                        if group == 1 and pd.notna(group_1_schedule):
                            return f"📚{pair_number}📚\n{pair_time}\n{group_1_schedule}\n"
                        elif group == 2 and pd.notna(group_2_schedule):
                            return f"📚{pair_number}📚\n{pair_time}\n{group_2_schedule}\n"
                except (ValueError, AttributeError):
                    continue

                # Если нашлись строки с практикой — добавляем строку "Практика в школе"
                if practice_counter > 0:
                    next_class = "Практика в школе.\n"
                break

    # Добавляем дату последнего обновления расписания
    update_time = get_last_update_time()
    schedule += f"\n📌 Дата последнего обновления: {update_time}"

    return next_class

# Главное меню
@router.message(Command(commands=['start']))
async def send_welcome(message: types.Message):
    markup = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="1 группа"), KeyboardButton(text="2 группа")]
        ],
        resize_keyboard=True
    )
    await message.answer("Привет! Выбери свою группу:", reply_markup=markup)

# Выбор группы
@router.message(lambda message: message.text in ["1 группа", "2 группа"])
async def choose_group(message: types.Message, state: FSMContext):
    group = 1 if message.text == "1 группа" else 2
    await state.update_data(group=group)
    markup = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сегодня"), KeyboardButton(text="Завтра")],
            [KeyboardButton(text="Неделя"), KeyboardButton(text="Выбрать дату")],
            [KeyboardButton(text="Следующая пара")]
        ],
        resize_keyboard=True
    )
    await message.answer(f"Вы выбрали {message.text}. Что показать?", reply_markup=markup)

# Показ расписания
@router.message(lambda message: message.text in ["Сегодня", "Завтра", "Неделя", "Выбрать дату", "Следующая пара"])
async def show_schedule(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("group", 1)
    today = datetime.now().date()
    
    if message.text == "Сегодня":
        date = today
        schedule = get_schedule(group, date)
        await message.answer(f"Расписание на сегодня ({date.strftime('%d.%m.%Y')}):\n\n{schedule}\n📌 Дата последнего обновления: {update_time}", parse_mode="HTML")
    elif message.text == "Завтра":
        date = today + timedelta(days=1)
        schedule = get_schedule(group, date)
        await message.answer(f"Расписание на завтра ({date.strftime('%d.%m.%Y')}):\n\n{schedule}\n📌 Дата последнего обновления: {update_time}")
    elif message.text == "Неделя":
        schedule = "Расписание на неделю:\n"
        for i in range(7):
            date = today + timedelta(days=i)
            daily_schedule = get_schedule(group, date)
            schedule += f"\n{date.strftime('%d.%m.%Y')}:\n{daily_schedule}"
        schedule += f"\n📌 Дата последнего обновления: {update_time}"
        await message.answer(schedule, parse_mode="HTML")
    elif message.text == "Выбрать дату":
        await message.answer("Введите дату в формате ДД.ММ.ГГГГ:")
    elif message.text == "Следующая пара":
        current_time = datetime.now().strftime("%H.%M")
        next_class = get_next_class(group, today, current_time)
        next_class += f"\n📌 Дата последнего обновления: {update_time}"
        await message.answer(next_class)
        

# Обработка ввода даты
@router.message(lambda message: bool(datetime.strptime(message.text, "%d.%m.%Y") if message.text else False))
async def custom_date_schedule(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("group", 1)
    try:
        date = datetime.strptime(message.text, "%d.%m.%Y").date()
        schedule = get_schedule(group, date)
        await message.answer(f"Расписание на {message.text}:\n\n{schedule}\n📌 Дата последнего обновления: {update_time}", parse_mode="HTML")
    except ValueError:
        await message.answer("Неверный формат даты. Попробуйте ещё раз.")

async def on_startup():
    if not check_update_flag():
        for user_id in subscribed_users:
            try:
                await bot.send_message(user_id, "⚙️ Бот был перезагружен. Нажмите /start для правильного отображения данных.")
            except Exception as e:
                logging.error(f"Ошибка отправки сообщения пользователю {user_id}: {e}")
    clear_update_flag()  # После отправки сообщения очищаем флаг

async def main():
    logging.info("Бот запущен и готов к работе!")
    dp.include_router(router)
    asyncio.create_task(auto_update())  # Фоновая проверка обновлений
    await on_startup()  # Отправка сообщения при запуске (если требуется)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
